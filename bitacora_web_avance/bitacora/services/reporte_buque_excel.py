import io
import os

from datetime import datetime

from django.conf import settings
from django.http import HttpResponse


def exportar_reporte_buques_excel(
    rows,
    fecha_inicio,
    fecha_fin,
    usuario_nombre="",
    usuario_cargo="",
):
    try:
        import openpyxl

        from openpyxl.styles import (
            Font,
            Alignment,
        )

    except ImportError as exc:

        raise RuntimeError(
            "La dependencia openpyxl no está instalada."
        ) from exc

    # ------------------------------------------
    # PLANTILLA
    # ------------------------------------------

    template_path = getattr(
        settings,
        "RUTA_PLANTILLA_REPORTE_BUQUES",
        "",
    )

    if not template_path:

        raise FileNotFoundError(
            "No se ha configurado "
            "RUTA_PLANTILLA_REPORTE_BUQUES."
        )

    template_path = os.fspath(
        template_path
    )

    if not os.path.exists(template_path):

        raise FileNotFoundError(
            "No se encontró la plantilla Excel: "
            f"{template_path}"
        )

    # ------------------------------------------
    # ABRIR PLANTILLA
    # ------------------------------------------

    workbook = openpyxl.load_workbook(
        template_path
    )

    worksheet = workbook.active

    # Si tu plantilla tiene una hoja específica,
    # cambia esto por el nombre correspondiente.
    #
    # Ejemplo:
    #
    # worksheet = workbook["Hoja1"]

    # ------------------------------------------
    # CABECERA
    # ------------------------------------------

    worksheet["A5"] = (
        "Fecha de Emisión : "
        f"{datetime.now().strftime('%d/%m/%Y')}"
    )

    worksheet["A6"] = (
        f"F.Desde : "
        f"{fecha_inicio.strftime('%d/%m/%Y')}"
        f"           "
        f"F.Hasta: "
        f"{fecha_fin.strftime('%d/%m/%Y')}"
    )

    # ------------------------------------------
    # FILA INICIAL
    # ------------------------------------------

    START_ROW = 9

    # ------------------------------------------
    # COLUMNAS
    # ------------------------------------------

    columnas = [
        "idregistro",
        "buque",
        "bandera",
        "scregistro",
        "tipo_de_trafico",
        "armador",
        "procedencia",
        "destino",
        "estado",
        "fecha_arribo",
        "fecha_zarpe",
    ]

    # ------------------------------------------
    # NORMALIZAR CLAVES
    # ------------------------------------------

    def normalizar_clave(valor):

        if valor is None:
            return ""

        return (
            str(valor)
            .strip()
            .lower()
            .replace(" ", "_")
        )

    # ------------------------------------------
    # PREPARAR REGISTRO
    # ------------------------------------------

    def preparar_registro(registro):

        resultado = {}

        for key, value in registro.items():

            clave = normalizar_clave(key)

            resultado[clave] = value

        return resultado

    # ------------------------------------------
    # ALIAS
    # ------------------------------------------

    aliases = {

        "idregistro": [
            "idregistro",
            "id_registro",
        ],

        "buque": [
            "buque",
            "nombre_buque",
        ],

        "bandera": [
            "bandera",
        ],

        "scregistro": [
            "scregistro",
            "sc_registro",
        ],

        "tipo_de_trafico": [
            "tipo_de_trafico",
            "tipo_trafico",
            "tiponave",
        ],

        "armador": [
            "armador",
        ],

        "procedencia": [
            "procedencia",
        ],

        "destino": [
            "destino",
        ],

        "estado": [
            "estado",
        ],

        "fecha_arribo": [
            "fecha_arribo",
            "fecha_arrivo",
            "f_arribo",
            "arribo",
        ],

        "fecha_zarpe": [
            "fecha_zarpe",
            "f_zarpe",
            "zarpe",
        ],
    }

    # ------------------------------------------
    # OBTENER VALOR
    # ------------------------------------------

    def obtener_valor(
        registro,
        campo,
    ):

        posibles = aliases.get(
            campo,
            [campo],
        )

        for posible in posibles:

            clave = normalizar_clave(
                posible
            )

            if clave in registro:

                valor = registro.get(
                    clave
                )

                if valor is not None:

                    return valor

        return ""

    # ------------------------------------------
    # ESCRIBIR REGISTROS
    # ------------------------------------------
    estilo_negro = Font(name="Calibri", size=11, color="000000", bold=False)

    for numero_fila, registro_original in enumerate(
        rows,
        start=START_ROW,
    ):

        registro = preparar_registro(
            registro_original
        )

        for numero_columna, campo in enumerate(
            columnas,
            start=1,
        ):

            valor = obtener_valor(
                registro,
                campo,
            )

            cell = worksheet.cell(
                row=numero_fila,
                column=numero_columna,
                value=valor,
            )
            cell.font = estilo_negro

    # ------------------------------------------
    # FIRMA
    # ------------------------------------------

    if rows:

        ultima_fila_datos = (
            START_ROW
            + len(rows)
            - 1
        )

    else:

        ultima_fila_datos = START_ROW

    sig_start = (
        ultima_fila_datos + 3
    )

    # PREPARADO POR

    worksheet.merge_cells(
        start_row=sig_start,
        start_column=2,
        end_row=sig_start,
        end_column=5,
    )

    cell_prep = worksheet.cell(
        row=sig_start,
        column=2,
    )

    cell_prep.value = "PREPARADO POR:"

    cell_prep.font = Font(
        name="Calibri",
        size=11,
    )

    cell_prep.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # Línea izquierda

    worksheet.merge_cells(
        start_row=sig_start + 2,
        start_column=2,
        end_row=sig_start + 2,
        end_column=5,
    )

    cell_line_left = worksheet.cell(
        row=sig_start + 2,
        column=2,
    )

    cell_line_left.value = (
        "________________________________________"
    )

    cell_line_left.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # Nombre

    worksheet.merge_cells(
        start_row=sig_start + 3,
        start_column=2,
        end_row=sig_start + 3,
        end_column=5,
    )

    cell_name_left = worksheet.cell(
        row=sig_start + 3,
        column=2,
    )

    cell_name_left.value = (
        usuario_nombre or ""
    )

    cell_name_left.font = Font(
        name="Calibri",
        size=11,
        bold=True,
    )

    cell_name_left.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # Cargo

    worksheet.merge_cells(
        start_row=sig_start + 4,
        start_column=2,
        end_row=sig_start + 4,
        end_column=5,
    )

    cell_cargo_left = worksheet.cell(
        row=sig_start + 4,
        column=2,
    )

    cell_cargo_left.value = (
        usuario_cargo or ""
    )

    cell_cargo_left.font = Font(
        name="Calibri",
        size=10,
    )

    cell_cargo_left.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # REVISADO

    worksheet.merge_cells(
        start_row=sig_start,
        start_column=9,
        end_row=sig_start,
        end_column=12,
    )

    cell_revisado = worksheet.cell(
        row=sig_start,
        column=9,
    )

    cell_revisado.value = "REVISADO"

    cell_revisado.font = Font(
        name="Calibri",
        size=11,
    )

    cell_revisado.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # Línea derecha

    worksheet.merge_cells(
        start_row=sig_start + 2,
        start_column=9,
        end_row=sig_start + 2,
        end_column=12,
    )

    cell_line_right = worksheet.cell(
        row=sig_start + 2,
        column=9,
    )

    cell_line_right.value = (
        "________________________________________"
    )

    cell_line_right.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # ------------------------------------------
    # GENERAR ARCHIVO
    # ------------------------------------------

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    filename = (
        "Reporte_Buques_"
        f"{fecha_inicio.strftime('%Y-%m-%d')}"
        "_a_"
        f"{fecha_fin.strftime('%Y-%m-%d')}.xlsx"
    )

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response