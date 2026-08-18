import logging
from datetime import datetime
from django.http import HttpResponse
from datetime import date
import io
import os

from django.conf import settings
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods

from .forms import LoginForm, RegistroCombustibleFilterForm
from .services import (
    DatabaseConfigurationError,
    DatabaseContractError,
    obtener_buques_artesanales,
    obtener_buques_industriales,
    obtener_reporte_combustible,
    obtener_turnos_usuario,
    validar_usuario,
    obtener_reporte_inec,
    obtener_tarifas_existentes,
    obtener_partidas,
    obtener_tasa_por_id,
    obtener_siguiente_codigo_tarifa,
    guardar_tarifa,
    anular_tarifa,
)

logger = logging.getLogger(__name__)


def _iniciar_sesion(request, datos_usuario, turnos):
    request.session.cycle_key()
    request.session["usuario_id"] = datos_usuario["idusuario"]
    request.session["usuario_login"] = datos_usuario["usuario"]
    request.session["usuario_nombre"] = datos_usuario["nombre"]
    request.session["usuario_cargo"] = turnos[0].get("cargo") or datos_usuario.get(
        "cargo",
        "Inspector",
    )


@never_cache
@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.session.get("usuario_id"):
        return redirect("bitacora_home")

    form = LoginForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            usuario = form.cleaned_data["usuario"]
            clave = form.cleaned_data["clave"]

            try:
                datos_usuario = validar_usuario(usuario, clave)

                if not datos_usuario:
                    messages.error(request, "Usuario o contraseña incorrectos.")
                else:

                    turnos = obtener_turnos_usuario(datos_usuario["idusuario"])
                    if not turnos:
                        messages.error(
                            request,
                            "El usuario es válido, pero no tiene un turno activo "
                            "habilitado para la bitácora.",
                        )
                    else:
                        _iniciar_sesion(request, datos_usuario, turnos)
                        return redirect("bitacora_home")

            except (DatabaseConfigurationError, DatabaseContractError) as exc:
                logger.exception("Configuración o contrato de base de datos inválido")
                messages.error(request, str(exc))
            except Exception:
                logger.exception("Error inesperado al validar el acceso")
                messages.error(
                    request,
                    "No fue posible comunicarse con la base de datos. "
                    "Revise la conexión o consulte al administrador.",
                )

    return render(
        request,
        "bitacora/login.html",
        {"demo_mode": settings.DEMO_MODE, "form": form},
    )


@never_cache
@require_http_methods(["GET"])
def bitacora_home(request):
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return redirect("login")

    try:
        turnos = obtener_turnos_usuario(idusuario)
        if not turnos:
            request.session.flush()
            messages.error(
                request,
                "Su turno ya no está activo o perdió el permiso de bitácora.",
            )
            return redirect("login")

        industriales = obtener_buques_industriales()
        artesanales = obtener_buques_artesanales()

    except (DatabaseConfigurationError, DatabaseContractError) as exc:
        logger.exception("Error de configuración al abrir la bitácora")
        messages.error(request, str(exc))
        turnos, industriales, artesanales = [], [], []
    except Exception:
        logger.exception("Error inesperado al cargar la bitácora")
        messages.error(
            request,
            "No fue posible cargar los datos de la bitácora desde SQL Server.",
        )
        turnos, industriales, artesanales = [], [], []

    return render(
        request,
        "bitacora/home.html",
        {
            "usuario_nombre": request.session.get("usuario_nombre"),
            "usuario_login": request.session.get("usuario_login"),
            "usuario_cargo": request.session.get("usuario_cargo"),
            "turnos": turnos,
            "buques_industriales": industriales,
            "buques_artesanales": artesanales,
            "demo_mode": settings.DEMO_MODE,
        },
    )


@never_cache
@require_http_methods(["GET"])
def tarifa_view(request):
    """Página base del tarifario."""
    if not request.session.get("usuario_id"):
        return redirect("login")

    return render(
        request,
        "bitacora/tarifa.html",
        {
            "usuario_nombre": request.session.get("usuario_nombre"),
            "usuario_login": request.session.get("usuario_login"),
            "usuario_cargo": request.session.get("usuario_cargo"),
            "demo_mode": settings.DEMO_MODE,
        },
    )

def exportar_reporte_inec_excel(
    rows,
    fecha_inicio,
    fecha_fin,
    usuario_nombre="",
    usuario_cargo="",
):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError as exc:
        raise RuntimeError(
            "La dependencia openpyxl no está instalada."
        ) from exc

    template_path = getattr(
        settings,
        "RUTA_PLANTILLA_INEC",
        "",
    )

    if not template_path:
        raise FileNotFoundError(
            "No se ha configurado RUTA_PLANTILLA_INEC."
        )

    template_path = os.fspath(template_path)

    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró la plantilla Excel: {template_path}"
        )

    workbook = openpyxl.load_workbook(
        template_path
    )

    if "Hoja3" in workbook.sheetnames:
        worksheet = workbook["Hoja3"]
    else:
        worksheet = workbook.active

    worksheet["A5"] = (
        f"Fecha de Emisión : "
        f"{datetime.now().strftime('%d/%m/%Y')}"
    )

    worksheet["A6"] = (
        f"F.Desde : {fecha_inicio.strftime('%d/%m/%Y')}"
        f"           "
        f"F.Hasta: {fecha_fin.strftime('%d/%m/%Y')}"
    )

    START_ROW = 9

    columnas = [
        "idregistro",
        "idbuqye",
        "matricula",
        "buque",
        "tipo_nave",
        "Trafico",
        "arribo",
        "zarpe",
        "bandera",
        "TRB",
        "TRN",
        "Pasajeros _e",
        "Pasajeros_s",
        "agencia",
        "Eslora",
        "Calado",
        "tipo_contrato",
        "p_bruto",
        "p_neto",
        "estado",
        "usuario",
        "scregistro",
        "mes",
        "tonelada",
    ]

    def normalizar_clave(valor):

        if valor is None:
            return ""

        return (
            str(valor)
            .strip()
            .lower()
            .replace(" ", "_")
        )

    def preparar_registro(registro):

        resultado = {}

        for key, value in registro.items():

            clave = normalizar_clave(key)

            resultado[clave] = value

        return resultado

    aliases = {
        "idregistro": [
            "idregistro",
            "id_registro",
        ],

        "idbuqye": [
            "idbuqye",
            "idbuque",
            "id_buque",
        ],

        "matricula": [
            "matricula",
            "matrícula",
        ],

        "buque": [
            "buque",
            "nombre_buque",
        ],

        "tipo_nave": [
            "tipo_nave",
            "tiponave",
            "tipo_navegacion",
        ],

        "Trafico": [
            "trafico",
            "tráfico",
        ],

        "arribo": [
            "arribo",
            "fecha_arribo",
        ],

        "zarpe": [
            "zarpe",
            "fecha_zarpe",
        ],

        "bandera": [
            "bandera",
        ],

        "TRB": [
            "trb",
        ],

        "TRN": [
            "trn",
        ],

        "Pasajeros _e": [
            "pasajeros_e",
            "pasajeros_entrada",
        ],

        "Pasajeros_s": [
            "pasajeros_s",
            "pasajeros_salida",
        ],

        "agencia": [
            "agencia",
            "agencia_naviera",
        ],

        "Eslora": [
            "eslora",
        ],

        "Calado": [
            "calado",
        ],

        "tipo_contrato": [
            "tipo_contrato",
            "tipocontrato",
        ],

        "p_bruto": [
            "p_bruto",
            "peso_bruto",
            "peso_bruto_total",
        ],

        "p_neto": [
            "p_neto",
            "peso_neto",
            "peso_neto_total",
        ],

        "estado": [
            "estado",
        ],

        "usuario": [
            "usuario",
            "usuario_registro",
        ],

        "scregistro": [
            "scregistro",
            "sc_registro",
        ],

        "mes": [
            "mes",
        ],

        "tonelada": [
            "tonelada",
            "toneladas",
        ],
    }

    def obtener_valor(registro, campo):

        posibles = aliases.get(
            campo,
            [campo],
        )

        for posible in posibles:

            clave = normalizar_clave(posible)

            if clave in registro:

                valor = registro.get(clave)

                if valor is not None:
                    return valor

        return ""

    for numero_fila, registro_original in enumerate(
        rows,
        start=START_ROW,
    ):

        registro = preparar_registro(
            registro_original
        )

        for numero_columna, campo in enumerate(
            columnas,
            start=1,
        ):

            valor = obtener_valor(
                registro,
                campo,
            )

            worksheet.cell(
                row=numero_fila,
                column=numero_columna,
                value=valor,
            )

    if rows:
        ultima_fila_datos = (
            START_ROW
            + len(rows)
            - 1
        )
    else:
        ultima_fila_datos = START_ROW

    sig_start = ultima_fila_datos + 3

    worksheet.merge_cells(
        start_row=sig_start,
        start_column=2,
        end_row=sig_start,
        end_column=5,
    )

    cell_prep = worksheet.cell(
        row=sig_start,
        column=2,
    )

    cell_prep.value = "PREPARADO POR:"

    cell_prep.font = Font(
        name="Calibri",
        size=11,
        bold=False,
    )

    cell_prep.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=sig_start + 2,
        start_column=2,
        end_row=sig_start + 2,
        end_column=5,
    )

    cell_line_left = worksheet.cell(
        row=sig_start + 2,
        column=2,
    )

    cell_line_left.value = (
        "________________________________________"
    )

    cell_line_left.font = Font(
        name="Calibri",
        size=11,
        bold=False,
    )

    cell_line_left.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=sig_start + 2,
        start_column=9,
        end_row=sig_start + 2,
        end_column=12,
    )

    cell_line_right = worksheet.cell(
        row=sig_start + 2,
        column=9,
    )

    cell_line_right.value = (
        "________________________________________"
    )

    cell_line_right.font = Font(
        name="Calibri",
        size=11,
        bold=False,
    )

    cell_line_right.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=sig_start + 3,
        start_column=2,
        end_row=sig_start + 3,
        end_column=5,
    )

    cell_name_left = worksheet.cell(
        row=sig_start + 3,
        column=2,
    )

    cell_name_left.value = usuario_nombre or ""

    cell_name_left.font = Font(
        name="Calibri",
        size=11,
        bold=True,
    )

    cell_name_left.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=sig_start + 4,
        start_column=2,
        end_row=sig_start + 4,
        end_column=5,
    )

    cell_cargo_left = worksheet.cell(
        row=sig_start + 4,
        column=2,
    )

    cell_cargo_left.value = usuario_cargo or ""

    cell_cargo_left.font = Font(
        name="Calibri",
        size=10,
        bold=False,
    )

    cell_cargo_left.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=sig_start,
        start_column=9,
        end_row=sig_start,
        end_column=12,
    )

    cell_revisado = worksheet.cell(
        row=sig_start,
        column=9,
    )

    cell_revisado.value = "REVISADO"

    cell_revisado.font = Font(
        name="Calibri",
        size=11,
        bold=False,
    )

    cell_revisado.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=sig_start + 3,
        start_column=9,
        end_row=sig_start + 3,
        end_column=12,
    )

    cell_name_right = worksheet.cell(
        row=sig_start + 3,
        column=9,
    )

    cell_name_right.value = "REVISADO"

    cell_name_right.font = Font(
        name="Calibri",
        size=11,
        bold=True,
    )

    cell_name_right.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    filename = (
        f"Reporte_INEC_"
        f"{fecha_inicio.strftime('%Y-%m-%d')}_"
        f"a_"
        f"{fecha_fin.strftime('%Y-%m-%d')}.xlsx"
    )

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response

@require_http_methods(["GET", "POST"])
def reporte_inec_view(request):

    if not request.session.get("usuario_id"):
        return redirect("login")

    contexto_base = {
        "usuario_nombre": request.session.get(
            "usuario_nombre",
            "",
        ),
        "usuario_login": request.session.get(
            "usuario_login",
            "",
        ),
        "usuario_cargo": request.session.get(
            "usuario_cargo",
            "",
        ),
        "demo_mode": settings.DEMO_MODE,
    }

    if request.method == "GET":

        return render(
            request,
            "bitacora/ReporteInec.html",
            contexto_base,
        )

    inicio = request.POST.get("inicio")
    fin = request.POST.get("fin")

    export = request.POST.get("export")

    try:

        fecha_inicio = datetime.strptime(
            inicio,
            "%Y-%m-%d",
        ).date()

        fecha_fin = datetime.strptime(
            fin,
            "%Y-%m-%d",
        ).date()

    except (
        TypeError,
        ValueError,
    ):

        messages.error(
            request,
            "Formato de fecha inválido.",
        )

        return render(
            request,
            "bitacora/ReporteInec.html",
            {
                **contexto_base,
                "inicio": inicio,
                "fin": fin,
            },
        )

    if fecha_inicio > fecha_fin:

        fecha_inicio, fecha_fin = (
            fecha_fin,
            fecha_inicio,
        )

    try:

        rows = obtener_reporte_inec(
            fecha_inicio,
            fecha_fin,
        )

    except (
        DatabaseConfigurationError,
        DatabaseContractError,
    ) as exc:

        logger.exception(
            "Error de configuración "
            "al obtener reporte INEC"
        )

        messages.error(
            request,
            str(exc),
        )

        rows = []

    except Exception:

        logger.exception(
            "Error inesperado "
            "al obtener reporte INEC"
        )

        messages.error(
            request,
            (
                "No fue posible obtener "
                "el reporte INEC desde SQL Server."
            ),
        )

        rows = []
    if export:

        if not rows:

            messages.warning(
                request,
                "No existen datos para exportar "
                "en el rango seleccionado.",
            )

            return render(
                request,
                "bitacora/ReporteInec.html",
                {
                    **contexto_base,
                    "rows": rows,
                    "inicio": fecha_inicio.isoformat(),
                    "fin": fecha_fin.isoformat(),
                },
            )

        try:

            return exportar_reporte_inec_excel(
                rows,
                fecha_inicio,
                fecha_fin,
                usuario_nombre=request.session.get(
                    "usuario_nombre",
                    "",
                ),
                usuario_cargo=request.session.get(
                    "usuario_cargo",
                    "",
                ),
            )

        except FileNotFoundError as exc:

            logger.exception(
                "No se encontró la plantilla INEC"
            )

            messages.error(
                request,
                str(exc),
            )

        except Exception:

            logger.exception(
                "Error al generar Excel INEC"
            )

            messages.error(
                request,
                (
                    "Ocurrió un error al generar "
                    "el archivo Excel."
                ),
            )

    return render(
        request,
        "bitacora/ReporteInec.html",
        {
            **contexto_base,
            "rows": rows,
            "inicio": fecha_inicio.isoformat(),
            "fin": fecha_fin.isoformat(),
        },
    )


@never_cache
@require_http_methods(["GET"])
def registro_combustible_home(request):
    if not request.session.get("usuario_id"):
        return redirect("login")

    form = RegistroCombustibleFilterForm(request.GET or None)
    registros: list[dict[str, object]] = []

    fecha_emision = date.today()
    fecha_desde = None
    fecha_hasta = None

    ajax_response = {
        "fecha_emision": fecha_emision.isoformat(),
        "fecha_desde": "",
        "fecha_hasta": "",
        "rows": [],
        "messages": [],
    }

    if form.is_valid():
        fecha_desde = form.cleaned_data.get("fecha_inicio")
        fecha_hasta = form.cleaned_data.get("fecha_fin")
        if fecha_desde:
            ajax_response["fecha_desde"] = fecha_desde.isoformat()
        if fecha_hasta:
            ajax_response["fecha_hasta"] = fecha_hasta.isoformat()

    if request.GET.get("buscar") == "1":
        if form.is_valid():
            try:
                registros = obtener_reporte_combustible(
                    form.cleaned_data["fecha_inicio"],
                    form.cleaned_data["fecha_fin"],
                )
                for registro in registros:
                    if "fecha_ingresa" in registro:
                        registro["fecha_ingresa"] = registro.get("fecha_ingresa", "")
                        registro["fecha"] = registro["fecha_ingresa"]
                    if "tikect" in registro:
                        registro["c_tikect"] = registro.get("tikect", "")
                        registro["c_ticket"] = registro["c_tikect"]
                    elif "c_tikect" in registro:
                        registro["c_ticket"] = registro.get("c_tikect", "")
                # Guardar último resultado de búsqueda en sesión para la exportación.
                try:
                    def _serialize_value(v):
                        if v is None:
                            return ""
                        # fechas y datetimes a ISO
                        if hasattr(v, "isoformat"):
                            try:
                                return v.isoformat()
                            except Exception:
                                pass
                        return str(v)

                    session_rows = [
                        {k: _serialize_value(v) for k, v in (reg.items())}
                        for reg in registros
                    ]
                    request.session["reporte_combustible_last"] = {
                        "fecha_desde": fecha_desde.isoformat() if fecha_desde else "",
                        "fecha_hasta": fecha_hasta.isoformat() if fecha_hasta else "",
                        "registros": session_rows,
                    }
                except Exception:
                    # No bloquear la búsqueda si la sesión no puede serializarse.
                    logger.exception("No fue posible guardar el resultado en sesión para exportación")
                if not registros:
                    msg = "No existen registros para el rango de fechas seleccionado."
                    messages.info(request, msg)
                    ajax_response["messages"].append({"text": msg, "tags": "info"})
            except (DatabaseConfigurationError, DatabaseContractError) as exc:
                logger.exception("Error de base de datos al obtener el reporte")
                generic_msg = "No fue posible obtener los datos del reporte. Revise la conexión o consulte al administrador."
                messages.error(request, generic_msg)
                ajax_response["messages"].append({"text": generic_msg, "tags": "error"})
            except Exception:
                logger.exception("Error inesperado al obtener el reporte de combustible")
                generic_msg = "No fue posible cargar los datos desde SQL Server. Revise la conexión o consulte al administrador."
                messages.error(request, generic_msg)
                ajax_response["messages"].append({"text": generic_msg, "tags": "error"})
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    ajax_response["messages"].append({"text": str(error), "tags": "error"})
            messages.error(
                request,
                "Complete correctamente Fecha Desde y Fecha Hasta antes de buscar.",
            )

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest" or "application/json" in request.headers.get("accept", "")

    if is_ajax:
        ajax_response["rows"] = [
            {
                "fecha_ingresa": str(registro.get("fecha_ingresa", registro.get("fecha", ""))),
                "c_tikect": str(registro.get("c_tikect", registro.get("c_ticket", registro.get("tikect", "")))),
                "guia": str(registro.get("guia", "")),
                "idplaca": str(registro.get("idplaca", "")),
                "chofer": str(registro.get("chofer", "")),
                "codbuque": str(registro.get("codbuque", "")),
                "buque": str(registro.get("buque", "")),
                "matricula": str(registro.get("matricula", "")),
                "galones": str(registro.get("galones", "")),
                "motivo": str(registro.get("motivo", "")),
            }
            for registro in registros
        ]
        return JsonResponse(ajax_response)

    return render(
        request,
        "bitacora/registro_combustible.html",
        {
            "demo_mode": settings.DEMO_MODE,
            "form": form,
            "registros": registros,
            "usuario_nombre": request.session.get("usuario_nombre", ""),
            "usuario_cargo": request.session.get("usuario_cargo", ""),
            "fecha_emision": fecha_emision,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
        },
    )


def _obtener_datos_exportacion(request):

    last = request.session.get("reporte_combustible_last")
    if last and last.get("registros"):
        return last

    form = RegistroCombustibleFilterForm(request.GET or None)
    if form.is_valid():
        try:
            fecha_inicio = form.cleaned_data["fecha_inicio"]
            fecha_fin = form.cleaned_data["fecha_fin"]
            registros = obtener_reporte_combustible(fecha_inicio, fecha_fin)
            for registro in registros:
                if "fecha_ingresa" in registro:
                    registro["fecha_ingresa"] = registro.get("fecha_ingresa", "")
                    registro["fecha"] = registro["fecha_ingresa"]
                if "tikect" in registro:
                    registro["c_tikect"] = registro.get("tikect", "")
                    registro["c_ticket"] = registro["c_tikect"]
                elif "c_tikect" in registro:
                    registro["c_ticket"] = registro.get("c_tikect", "")

            def _serialize_value(v):
                if v is None:
                    return ""
                if hasattr(v, "isoformat"):
                    try:
                        return v.isoformat()
                    except Exception:
                        pass
                return str(v)

            session_rows = [
                {k: _serialize_value(v) for k, v in reg.items()}
                for reg in registros
            ]
            datos = {
                "fecha_desde": fecha_inicio.isoformat(),
                "fecha_hasta": fecha_fin.isoformat(),
                "registros": session_rows,
            }
            request.session["reporte_combustible_last"] = datos
            return datos
        except Exception:
            logger.exception("Error al recuperar registros de exportación vía parámetros GET")
            pass

    return last


@require_http_methods(["GET"])
def exportar_excel(request):

    if not request.session.get("usuario_id"):
        return redirect("login")

    last = _obtener_datos_exportacion(request)

    if not last:
        messages.error(
            request,
            "Primero debe realizar una búsqueda para exportar la información."
        )
        return redirect("registro_combustible")

    registros = last.get("registros", [])

    if not registros:
        messages.info(
            request,
            "No existen registros para exportar."
        )
        return redirect("registro_combustible")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        messages.error(
            request,
            "La dependencia 'openpyxl' para generar Excel no está instalada."
        )
        return redirect("registro_combustible")

    template_path = getattr(
        settings,
        "RUTA_PLANTILLA_COMB",
        "",
    )

    if not template_path:
        raise FileNotFoundError(
            "No se ha configurado RUTA_PLANTILLA_COMB."
        )

    template_path = os.fspath(template_path)

    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró la plantilla Excel: {template_path}"
        )

    try:
        with open(template_path, "rb") as f:
            template_bytes = io.BytesIO(f.read())

        wb = openpyxl.load_workbook(template_bytes)

        if "controlcombustible" in wb.sheetnames:
            ws = wb["controlcombustible"]
        elif "Hoja2" in wb.sheetnames:
            ws = wb["Hoja2"]
        else:
            ws = wb.active

        def safe_write_cell(ws, r, c, val):
            cell = ws.cell(row=r, column=c)

            if cell.__class__.__name__ != "MergedCell":
                cell.value = val

        fecha_emision_str = date.today().strftime("%d/%m/%Y")

        safe_write_cell(
            ws,
            4,
            1,
            f"Fecha de Emisión : Manta, {fecha_emision_str}"
        )

        def _fmt_fecha(val):
            if not val:
                return "—"

            if hasattr(val, "strftime"):
                return val.strftime("%d/%m/%Y")

            val_str = str(val).split("T")[0]
            parts = val_str.split("-")

            if len(parts) == 3 and len(parts[0]) == 4:
                return f"{parts[2]}/{parts[1]}/{parts[0]}"

            return str(val)

        f_desde_str = _fmt_fecha(
            last.get("fecha_desde")
        )

        f_hasta_str = _fmt_fecha(
            last.get("fecha_hasta")
        )

        safe_write_cell(
            ws,
            5,
            1,
            (
                f"F.Desde: {f_desde_str}"
                f"               "
                f"F.Hasta :      {f_hasta_str}"
            ),
        )

        def first_value(row, keys):

            if isinstance(row, dict):

                for key in keys:
                    val = row.get(key)

                    if (
                        val is not None
                        and str(val).strip() != ""
                    ):
                        return val

            elif isinstance(row, (list, tuple)):

                for val in row:

                    if (
                        val is not None
                        and str(val).strip() != ""
                    ):
                        return val

            return ""

        for rng in list(ws.merged_cells.ranges):

            if rng.min_row >= 8:
                ws.unmerge_cells(str(rng))

        start_row = 8

        for idx, reg in enumerate(
            registros,
            start=start_row,
        ):

            fila = [
                first_value(
                    reg,
                    (
                        "fecha_ingresa",
                        "fecha",
                        "fecha_registro",
                        "fecha_mov",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "c_tikect",
                        "c_ticket",
                        "tikect",
                        "ticket",
                        "tickets",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "guia",
                        "guia_r",
                        "guia_no",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "idplaca",
                        "placa",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "chofer",
                        "conductor",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "licencia",
                        "licencia_conductor",
                        "licencia_no",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "codbuque",
                        "cod_buque",
                        "scbuque",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "buque",
                        "nombre",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "matricula",
                        "n_matricula",
                    ),
                ),

                first_value(
                    reg,
                    (
                        "galones",
                        "cantidad_litros",
                        "litros",
                    ),
                ),

                first_value(
                    reg,
                    ("motivo",),
                ),

                first_value(
                    reg,
                    ("estado",),
                ),

                first_value(
                    reg,
                    (
                        "tipo_carro",
                        "tipo carro",
                        "tipo",
                    ),
                ),
            ]

            for col_idx, value in enumerate(
                fila,
                start=1,
            ):
                safe_write_cell(
                    ws,
                    idx,
                    col_idx,
                    value,
                )

        last_data_row = (
            start_row
            + len(registros)
            - 1
        )

        sig_start = last_data_row + 3

        ws.merge_cells(
            start_row=sig_start,
            start_column=2,
            end_row=sig_start,
            end_column=5,
        )

        cell_prep = ws.cell(
            row=sig_start,
            column=2,
        )

        cell_prep.value = "PREPARADO POR:"
        cell_prep.font = Font(
            name="Calibri",
            size=11,
            bold=False,
        )
        cell_prep.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws.merge_cells(
            start_row=sig_start + 2,
            start_column=2,
            end_row=sig_start + 2,
            end_column=5,
        )

        cell_line_left = ws.cell(
            row=sig_start + 2,
            column=2,
        )

        cell_line_left.value = (
            "________________________________________"
        )

        cell_line_left.font = Font(
            name="Calibri",
            size=11,
            bold=False,
        )

        cell_line_left.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # Línea derecha
        ws.merge_cells(
            start_row=sig_start + 2,
            start_column=9,
            end_row=sig_start + 2,
            end_column=12,
        )

        cell_line_right = ws.cell(
            row=sig_start + 2,
            column=9,
        )

        cell_line_right.value = (
            "________________________________________"
        )

        cell_line_right.font = Font(
            name="Calibri",
            size=11,
            bold=False,
        )

        cell_line_right.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        #Nombre
        ws.merge_cells(
            start_row=sig_start + 3,
            start_column=2,
            end_row=sig_start + 3,
            end_column=5,
        )

        cell_name_left = ws.cell(
            row=sig_start + 3,
            column=2,
        )

        cell_name_left.value = request.session.get(
            "usuario_nombre",
            "",
        )

        cell_name_left.font = Font(
            name="Calibri",
            size=11,
            bold=True,
        )

        cell_name_left.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # cargo
        ws.merge_cells(
            start_row=sig_start + 4,
            start_column=2,
            end_row=sig_start + 4,
            end_column=5,
        )

        cell_cargo_left = ws.cell(
            row=sig_start + 4,
            column=2,
        )

        cell_cargo_left.value = request.session.get(
            "usuario_cargo",
            "",
        )

        cell_cargo_left.font = Font(
            name="Calibri",
            size=10,
            bold=False,
        )

        cell_cargo_left.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        
        # revisado
        ws.merge_cells(
            start_row=sig_start + 3,
            start_column=9,
            end_row=sig_start + 3,
            end_column=12,
        )

        cell_name_right = ws.cell(
            row=sig_start + 3,
            column=9,
        )

        cell_name_right.value = "REVISADO"

        cell_name_right.font = Font(
            name="Calibri",
            size=11,
            bold=True,
        )

        cell_name_right.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        for rng in list(ws.merged_cells.ranges):

            if rng.min_row > sig_start + 4:
                ws.unmerge_cells(str(rng))

        max_r = ws.max_row

        if max_r > sig_start + 4:

            ws.delete_rows(
                sig_start + 5,
                max_r - (sig_start + 4),
            )
        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        clean_desde = (
            f_desde_str
            .replace("/", "-")
            .replace("—", "sin_fecha")
            .strip()
        )

        clean_hasta = (
            f_hasta_str
            .replace("/", "-")
            .replace("—", "sin_fecha")
            .strip()
        )

        raw_filename = (
            f"ReporteCombustible_"
            f"{clean_desde}_"
            f"{clean_hasta}.xlsx"
        )

        filename_ascii = "".join(
            c
            if c.isalnum() or c in "._-"
            else "_"
            for c in raw_filename
        )

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename_ascii}"'
        )

        return response

    except Exception as exc:

        logger.exception(
            "Error al generar el archivo Excel"
        )

        messages.error(
            request,
            (
                "Ocurrió un error al generar "
                f"el archivo Excel: {exc}"
            ),
        )

        return redirect(
            "registro_combustible"
        )

@require_http_methods(["GET"])
def exportar_excel_validar(request):
    """Valida via AJAX si la exportación puede ejecutarse."""

    if not request.session.get("usuario_id"):
        return JsonResponse({
            "ok": False,
            "message": "Debe iniciar sesión para exportar.",
            "level": "error",
        })

    last = _obtener_datos_exportacion(request)

    if not last:
        return JsonResponse({
            "ok": False,
            "message": "Primero debe realizar una búsqueda para exportar la información.",
            "level": "info",
        })

    registros = last.get("registros", [])

    if not registros:
        return JsonResponse({
            "ok": False,
            "message": "No existen registros para exportar.",
            "level": "info",
        })

    template_path = getattr(
        settings,
        "RUTA_PLANTILLA_COMB",
        "",
    )

    if not template_path:
        return JsonResponse({
            "ok": False,
            "message": "No se ha configurado RUTA_PLANTILLA_COMB.",
            "level": "error",
        })

    template_path = os.fspath(template_path)

    if not os.path.exists(template_path):
        return JsonResponse({
            "ok": False,
            "message": (
                f"No se encontró la plantilla Excel: "
                f"{template_path}"
            ),
            "level": "error",
        })

    return JsonResponse({
        "ok": True,
    })


@require_http_methods(["POST"])
def logout_view(request):
    request.session.flush()
    messages.success(request, "Sesión cerrada correctamente.")
    return redirect("login")


@never_cache
@require_http_methods(["GET"])
def tarifa_view(request):
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return redirect("login")

    return render(
        request,
        "bitacora/tarifa.html",
        {
            "usuario_nombre": request.session.get("usuario_nombre"),
            "usuario_login": request.session.get("usuario_login"),
            "usuario_cargo": request.session.get("usuario_cargo"),
            "demo_mode": settings.DEMO_MODE,
        },
    )


@never_cache
@require_http_methods(["GET"])
def tarifa_inflacion_view(request):
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return redirect("login")

    try:
        tarifas = obtener_tarifas_existentes(estado=101)
    except Exception as exc:
        logger.exception("Error al obtener tarifas para inflación")
        tarifas = []

    return render(
        request,
        "bitacora/tarifa_inflacion.html",
        {
            "usuario_nombre": request.session.get("usuario_nombre"),
            "usuario_login": request.session.get("usuario_login"),
            "usuario_cargo": request.session.get("usuario_cargo"),
            "demo_mode": settings.DEMO_MODE,
            "tarifas": tarifas,
        },
    )


@never_cache
@require_http_methods(["GET"])
def tarifa_listado_view(request):
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return redirect("login")

    filtro = request.GET.get("filtro", "activas").strip().lower()
    estado = 100 if filtro == "todas" else 1

    try:
        tarifas = obtener_tarifas_existentes(estado=estado)
    except Exception as exc:
        logger.exception("Error al obtener tarifas existentes")
        tarifas = []

    return render(
        request,
        "bitacora/tarifa_listado.html",
        {
            "tarifas": tarifas,
            "filtro": filtro,
        },
    )

from django.http import JsonResponse, HttpResponse


@never_cache
@require_http_methods(["GET"])
def api_buscar_partida(request):
    """API Endpoint para consultar partidas mediante AJAX."""
    if not request.session.get("usuario_id"):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=401)

    codigo = request.GET.get("codigo", "1").strip()

    try:
        resultados = obtener_partidas(codigo)
        return JsonResponse({"success": True, "data": resultados})
    except (DatabaseConfigurationError, DatabaseContractError) as exc:
        logger.exception("Error de base de datos al buscar partida")
        return JsonResponse({"success": False, "error": str(exc)}, status=400)
    except Exception:
        logger.exception("Error inesperado al buscar partida")
        return JsonResponse(
            {"success": False, "error": "No se pudo consultar la partida."}, status=500
        )


@never_cache
@require_http_methods(["GET"])
def api_buscar_tasa(request):
    """API Endpoint para consultar tasas mediante AJAX."""
    if not request.session.get("usuario_id"):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=401)

    idtasa = request.GET.get("idtasa", "").strip()

    try:
        resultados = obtener_tasa_por_id(idtasa)
        return JsonResponse({"success": True, "data": resultados})
    except (DatabaseConfigurationError, DatabaseContractError) as exc:
        logger.exception("Error de base de datos al buscar tasa")
        return JsonResponse({"success": False, "error": str(exc)}, status=400)
    except Exception:
        logger.exception("Error inesperado al buscar tasa")
        return JsonResponse(
            {"success": False, "error": "No se pudo consultar la tasa."}, status=500
        )


@never_cache
@require_http_methods(["GET"])
def api_siguiente_codigo(request):
    """API Endpoint para consultar el siguiente código secuencial para una tasa."""
    if not request.session.get("usuario_id"):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=401)

    idtasa = request.GET.get("idtasa", "").strip()
    if not idtasa:
        return JsonResponse({"success": False, "error": "Falta idtasa"}, status=400)

    try:
        siguiente = obtener_siguiente_codigo_tarifa(idtasa)
        return JsonResponse({"success": True, "siguiente": siguiente})
    except Exception:
        logger.exception("Error al calcular el siguiente código de tarifa")
        return JsonResponse(
            {"success": False, "error": "No se pudo calcular el siguiente código."}, status=500
        )


@never_cache
@require_http_methods(["POST"])
def guardar_tarifa_view(request):
    """API Endpoint para guardar una tarifa nueva utilizando el SPJ_insert_Tarifas."""
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return JsonResponse({"success": False, "error": "No autorizado"}, status=401)

    codigo = request.POST.get("codigo", "").strip()
    tarifa = request.POST.get("tarifa", "").strip()
    valor = request.POST.get("valor", "0.0000").strip()
    partida_cod = request.POST.get("partida_cod", "").strip()
    partida_id = request.POST.get("partida_id", "").strip()
    tasa_id = request.POST.get("tasa_id", "").strip()
    formula = request.POST.get("formula", "").strip()
    detalle = request.POST.get("detalle", "").strip()
    calc_unidad = request.POST.get("calc_unidad", "").strip()
    calc_param = request.POST.get("calc_param", "").strip()
    iva = request.POST.get("iva", "0").strip()
    ticket_srv = request.POST.get("ticket_srv", "").strip()
    activa = request.POST.get("activa", "1").strip()
    permitir_cambio_valor = request.POST.get("permitir_cambio_valor", "0").strip()
    aplica_inflacion = request.POST.get("aplica_inflacion", "0").strip()
    idtarifa_raw = request.POST.get("id", "0").strip()
    try:
        idtarifa = int(idtarifa_raw)
    except (ValueError, TypeError):
        idtarifa = 0

    if not codigo or not tarifa or not tasa_id:
        return JsonResponse({"success": False, "error": "Faltan campos obligatorios (Código, Tarifa o Tasa)"})

    if len(codigo) > 5:
        return JsonResponse({"success": False, "error": "El Código no puede superar los 5 caracteres."})

    if len(tarifa) > 80:
        return JsonResponse({"success": False, "error": "La Tarifa no puede superar los 80 caracteres."})

    if len(formula) > 50:
        return JsonResponse({"success": False, "error": "La Fórmula no puede superar los 50 caracteres."})

    if len(detalle) > 50:
        return JsonResponse({"success": False, "error": "El Detalle no puede superar los 50 caracteres."})

    # Mapeo de parámetros del frontend a enteros esperados por el SP
    calc_unidad_map = {"dia": 1, "horas": 2}
    hora_dia = calc_unidad_map.get(calc_unidad, 3) # default a 3 (cantidad/otros)

    calc_param_map = {"eslora": 1, "t_neto": 2}
    eslora_tneto = calc_param_map.get(calc_param, 3) # default a 3 (otros)

    ticket_srv_map = {"vehiculo": 1, "muelle": 2}
    ticket = ticket_srv_map.get(ticket_srv, 0) # default a 0 (ninguno)

    try:
        resul = guardar_tarifa(
            idtarifa=idtarifa,
            codigo=codigo,
            tarifa=tarifa,
            valor=valor,
            partida_cod=partida_cod,
            partida_id=partida_id,
            tasa_id=tasa_id,
            formula=formula,
            detalle=detalle,
            hora_dia=hora_dia,
            eslora_tneto=eslora_tneto,
            iva=1 if iva in ["1", "true", "True"] else 0,
            ticket=ticket,
            activo=1 if activa in ["1", "true", "True"] else 0,
            cambio_factura=1 if permitir_cambio_valor in ["1", "true", "True"] else 0,
            aplica_inflacion=1 if aplica_inflacion in ["1", "true", "True"] else 0,
        )
        if resul == 3:
            return JsonResponse({"success": False, "error": "El código de tarifa ya existe para esta tasa."})
        elif resul == 4:
            return JsonResponse({"success": False, "error": "El nombre de tarifa ya existe para esta tasa o hay conflicto."})
        elif resul == -1:
            return JsonResponse({"success": False, "error": "Error interno en la base de datos al guardar la tarifa."})

        return JsonResponse({"success": True, "resul": resul})
    except Exception as exc:
        logger.exception("Error al ejecutar guardado de tarifa")
        return JsonResponse({"success": False, "error": str(exc)}, status=500)


@never_cache
@require_http_methods(["POST"])
def anular_tarifa_view(request):
    """API Endpoint para anular una tarifa estableciendo idestado = 7 y activo = 0."""
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return JsonResponse({"success": False, "error": "No autorizado"}, status=401)

    idtarifa = request.POST.get("id", "").strip()
    if not idtarifa or idtarifa == "0":
        return JsonResponse({"success": False, "error": "Debe seleccionar una tarifa guardada para poder anularla."})

    try:
        anular_tarifa(idtarifa)
        return JsonResponse({"success": True, "message": f"Tarifa con ID {idtarifa} anulada correctamente."})
    except Exception as exc:
        logger.exception("Error al anular tarifa")
        return JsonResponse({"success": False, "error": str(exc)}, status=500)


@never_cache
@require_http_methods(["GET"])
def exportar_tarifas_view(request):
    """Genera un archivo Excel con el listado de tarifas cargando la plantilla excel/Tarifas_J.xlsx."""
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from copy import copy
    
    idusuario = request.session.get("usuario_id")
    if not idusuario:
        return HttpResponse("No autorizado", status=401)

    filtro = request.GET.get("filtro", "activas").strip().lower()
    estado = 100 if filtro == "todas" else 1
    q = request.GET.get("q", "").strip().lower()

    try:
        # 1. Obtener tarifas según filtro
        tarifas = obtener_tarifas_existentes(estado=estado)

        # 1.5. Filtrar por término de búsqueda (búsqueda en vivo de la tabla)
        if q:
            tarifas = [
                t for t in tarifas
                if q in (t.get("tasa") or "").lower()
                or q in (t.get("codigo") or "").lower()
                or q in (t.get("tarifa") or "").lower()
            ]

        # Ordenar por el nombre de la tasa (tasa) y luego por codigo (COD.TARIFA) en orden ascendente
        def sort_key(t):
            tasa_nombre = (t.get("tasa") or "").strip().upper()
            codigo = t.get("codigo") or "0"
            try:
                cod_val = int(codigo)
            except ValueError:
                cod_val = codigo

            cod_key = (0, cod_val) if isinstance(cod_val, int) else (1, str(cod_val))
            return (tasa_nombre, cod_key)

        tarifas = sorted(tarifas, key=sort_key)

        # 2. Ruta a la plantilla
        template_path = getattr(
            settings,
            "RUTA_PLANTILLA_TARI",
            "",
        )

        if not template_path:
            raise FileNotFoundError(
                "No se ha configurado RUTA_PLANTILLA_TARI."
            )

        template_path = os.fspath(template_path)

        if not os.path.exists(template_path):
            raise FileNotFoundError(
                f"No se encontró la plantilla Excel: {template_path}"
        )

        # 3. Cargar el libro y la hoja activa
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Descombinar celdas combinadas de la fila 6 en adelante para evitar errores de escritura
        for r in list(ws.merged_cells.ranges):
            if r.min_row >= 6:
                ws.unmerge_cells(str(r))

        # Escribir la fecha de emisión en la celda A3
        from datetime import datetime
        current_date_str = datetime.now().strftime("%d/%m/%Y")
        ws.cell(row=5, column=1, value=f"Fecha de Emisión : {current_date_str}")
        ws.cell(row=6, column=1, value=None)  # Limpiar A4 por si acaso

        # 4. Escribir los datos en el excel a partir de la fila 8
        # Las columnas correspondientes son:
        # Col 1: Tasa (tasa)
        # Col 2: Cod.tarifa (codigo / sctarifa)
        # Col 3: Tarifa (tarifa)
        # Col 4: Valor (valor)
        # Col 5: Formula (formula)
        # Col 6: Inflacion (aplica_inflacion_txt)
        # Col 7: Detalle (detalle)
        # Col 8: Cod.partida (partida_cod / scpartida)
        # Col 9: Partida (partida_desc)
        start_row = 8
        for idx, t in enumerate(tarifas):
            row_num = start_row + idx
            
            try:
                val_num = float(t.get("valor", "0.0000"))
            except (ValueError, TypeError):
                val_num = 0.00

            try:
                cod_num = int(t.get("codigo", ""))
            except (ValueError, TypeError):
                cod_num = t.get("codigo", "")

            ws.cell(row=row_num, column=1, value=t.get("tasa", ""))
            ws.cell(row=row_num, column=2, value=cod_num)
            ws.cell(row=row_num, column=3, value=t.get("tarifa", ""))
            ws.cell(row=row_num, column=4, value=val_num)
            ws.cell(row=row_num, column=5, value=t.get("formula", ""))
            ws.cell(row=row_num, column=6, value=t.get("aplica_inflacion_txt", "No aplica Inflación anual"))
            ws.cell(row=row_num, column=7, value=t.get("detalle", ""))
            ws.cell(row=row_num, column=8, value=t.get("partida_cod", ""))
            ws.cell(row=row_num, column=9, value=t.get("partida_desc", ""))
            ws.cell(row=row_num, column=10, value="ACTIVA" if t.get("activa") else "ANULADA")

            # Formatear celdas de la fila y forzar color de texto negro
            for col_idx in range(1, 11):
                cell = ws.cell(row=row_num, column=col_idx)
                
                # Alternar colores copiando el formato de la fila 8 (azul oscuro) o fila 9 (azul claro) de la plantilla
                src_row = 8 if row_num % 2 == 0 else 9
                src_cell = ws.cell(row=src_row, column=col_idx)
                
                if src_cell.has_style:
                    cell.font = copy(src_cell.font)
                    cell.border = copy(src_cell.border)
                    cell.fill = copy(src_cell.fill)
                    cell.number_format = copy(src_cell.number_format)
                    cell.alignment = copy(src_cell.alignment)

                # Forzar color de texto negro y desactivar negritas para legibilidad
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name or "Calibri",
                        size=cell.font.size or 11,
                        bold=False,  # Desactivar negrita heredada de la plantilla
                        italic=cell.font.italic or False,
                        color="000000",  # Negro
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        vertAlign=cell.font.vertAlign,
                        scheme=cell.font.scheme
                    )
                else:
                    cell.font = Font(name="Calibri", size=11, color="000000")

        # 4.5. Agregar bloque de firmas al final de todas las tarifas
        last_data_row = start_row + len(tarifas) - 1 if len(tarifas) > 0 else 7
        sig_start = last_data_row + 3

        # Escribir "PREPARADO POR:" en la columna 1 (A)
        cell_prep = ws.cell(row=sig_start, column=1)
        cell_prep.value = "PREPARADO POR:"
        cell_prep.font = Font(name="Calibri", size=11, bold=True)
        cell_prep.alignment = Alignment(horizontal="left", vertical="center")

        # Línea de firma izquierda (Columnas 2 a 4)
        ws.merge_cells(
            start_row=sig_start + 2,
            start_column=2,
            end_row=sig_start + 2,
            end_column=4,
        )
        cell_line_left = ws.cell(row=sig_start + 2, column=2)
        cell_line_left.value = "________________________________________"
        cell_line_left.font = Font(name="Calibri", size=11, bold=False)
        cell_line_left.alignment = Alignment(horizontal="center", vertical="center")

        # Nombre del usuario que inicia sesión (Columnas 2 a 4)
        ws.merge_cells(
            start_row=sig_start + 3,
            start_column=2,
            end_row=sig_start + 3,
            end_column=4,
        )
        cell_name_left = ws.cell(row=sig_start + 3, column=2)
        cell_name_left.value = request.session.get("usuario_nombre", "")
        cell_name_left.font = Font(name="Calibri", size=11, bold=True)
        cell_name_left.alignment = Alignment(horizontal="center", vertical="center")

        # Cargo del usuario que inicia sesión (Columnas 2 a 4)
        ws.merge_cells(
            start_row=sig_start + 4,
            start_column=2,
            end_row=sig_start + 4,
            end_column=4,
        )
        cell_cargo_left = ws.cell(row=sig_start + 4, column=2)
        cell_cargo_left.value = request.session.get("usuario_cargo", "")
        cell_cargo_left.font = Font(name="Calibri", size=10, bold=False)
        cell_cargo_left.alignment = Alignment(horizontal="center", vertical="center")

        # Línea de firma derecha "REVISADO" (Columnas 7 a 9)
        ws.merge_cells(
            start_row=sig_start + 2,
            start_column=7,
            end_row=sig_start + 2,
            end_column=9,
        )
        cell_line_right = ws.cell(row=sig_start + 2, column=7)
        cell_line_right.value = "________________________________________"
        cell_line_right.font = Font(name="Calibri", size=11, bold=False)
        cell_line_right.alignment = Alignment(horizontal="center", vertical="center")

        # Texto "REVISADO" (Columnas 7 a 9)
        ws.merge_cells(
            start_row=sig_start + 3,
            start_column=7,
            end_row=sig_start + 3,
            end_column=9,
        )
        cell_title_right = ws.cell(row=sig_start + 3, column=7)
        cell_title_right.value = "REVISADO"
        cell_title_right.font = Font(name="Calibri", size=11, bold=True)
        cell_title_right.alignment = Alignment(horizontal="center", vertical="center")

        # 5. Generar respuesta HTTP para la descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Tarifas_Reporte.xlsx"'
        wb.save(response)
        return response

    except Exception as exc:
        logger.exception("Error al exportar tarifas a excel")
        return HttpResponse(f"Error interno al exportar Excel: {str(exc)}", status=500)